"""General customers CRUD + import."""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any, List
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from pydantic import BaseModel, Field
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.auth.deps import require_any_permission, require_permission
from app.db import get_db
from app.models.general_customer import GeneralCustomer as GeneralCustomerModel
from app.services.audit_service import (
    ACTION_CREATE,
    ACTION_DELETE,
    ACTION_UPDATE,
    get_client_ip,
    log_action,
)

router = APIRouter()


class GeneralCustomerOut(BaseModel):
    id: UUID
    customer_id: str
    customer_name: str | None
    created_at: datetime


class GeneralCustomerCreate(BaseModel):
    customer_id: str = Field(..., min_length=1, max_length=64)
    customer_name: str | None = Field(default=None, max_length=255)


class GeneralCustomerUpdate(BaseModel):
    customer_name: str | None = Field(default=None, max_length=255)


class GeneralCustomerImportError(BaseModel):
    row: int
    detail: str


class GeneralCustomerImportResult(BaseModel):
    created: int
    updated: int
    errors: list[GeneralCustomerImportError] = Field(default_factory=list)


def _to_out(item: GeneralCustomerModel) -> GeneralCustomerOut:
    return GeneralCustomerOut(
        id=item.id,
        customer_id=item.customer_id,
        customer_name=item.customer_name,
        created_at=item.created_at,
    )


def _norm_header_key(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "_").replace("-", "_")


def _row_dict_from_csv(raw: bytes) -> tuple[list[dict[str, str]], list[GeneralCustomerImportError]]:
    errors: list[GeneralCustomerImportError] = []
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        errors.append(GeneralCustomerImportError(row=0, detail="File must be UTF-8 CSV"))
        return [], errors
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        errors.append(GeneralCustomerImportError(row=0, detail="CSV has no header row"))
        return [], errors
    rows: list[dict[str, str]] = []
    for row in reader:
        clean: dict[str, str] = {}
        for k, v in row.items():
            if k is None:
                continue
            nk = _norm_header_key(str(k))
            if nk:
                clean[nk] = (str(v).strip() if v is not None else "")
        rows.append(clean)
    return rows, errors


def _row_dicts_from_xlsx(raw: bytes) -> tuple[list[dict[str, str]], list[GeneralCustomerImportError]]:
    errors: list[GeneralCustomerImportError] = []
    try:
        from openpyxl import load_workbook
    except ImportError:
        errors.append(GeneralCustomerImportError(row=0, detail="xlsx support not available (openpyxl)"))
        return [], errors
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        errors.append(GeneralCustomerImportError(row=0, detail=f"Invalid xlsx: {e}"))
        return [], errors
    ws = wb.active
    data_rows: list[tuple[Any, ...]] = []
    for row in ws.iter_rows(values_only=True):
        data_rows.append(row)
    wb.close()
    if not data_rows:
        errors.append(GeneralCustomerImportError(row=0, detail="Empty sheet"))
        return [], errors
    headers_raw = data_rows[0]
    headers = [_norm_header_key(str(h) if h is not None else "") for h in (headers_raw or ())]
    if not any(h for h in headers):
        errors.append(GeneralCustomerImportError(row=0, detail="Missing header row"))
        return [], errors
    rows: list[dict[str, str]] = []
    for vals in data_rows[1:]:
        row: dict[str, str] = {}
        for j, h in enumerate(headers):
            if not h:
                continue
            cell = vals[j] if j < len(vals) else None
            row[h] = str(cell).strip() if cell is not None else ""
        rows.append(row)
    return rows, errors


def _get_cell(row: dict[str, str], *keys: str) -> str:
    for k in keys:
        nk = _norm_header_key(k)
        if nk in row:
            return (row.get(nk) or "").strip()
    return ""


@router.get("", response_model=List[GeneralCustomerOut], summary="List general customers")
@router.get("/", response_model=List[GeneralCustomerOut], summary="List general customers")
async def list_general_customers(
    search: str | None = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    _user=Depends(require_any_permission(["orders:read", "receiving:write", "admin:access"])),
):
    query = db.query(GeneralCustomerModel)
    if search:
        term = f"%{search.strip()}%"
        query = query.filter(
            or_(
                GeneralCustomerModel.customer_id.ilike(term),
                GeneralCustomerModel.customer_name.ilike(term),
            )
        )
    items = query.order_by(GeneralCustomerModel.customer_id.asc()).offset(offset).limit(limit).all()
    return [_to_out(i) for i in items]


@router.post(
    "/import",
    response_model=GeneralCustomerImportResult,
    summary="Import general customers from CSV or xlsx (upsert by customer_id)",
)
async def import_general_customers(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(require_permission("orders:read")),
):
    raw = await file.read()
    if len(raw) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 8MB)")
    name = (file.filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        rows, errors = _row_dicts_from_xlsx(raw)
    else:
        rows, errors = _row_dict_from_csv(raw)
    if not rows and not errors:
        raise HTTPException(status_code=400, detail="No data rows found")

    created = 0
    updated = 0
    for i, row in enumerate(rows, start=2):
        cid = _get_cell(row, "customer_id", "customerid", "mijoz_id", "id")
        cname = _get_cell(row, "customer_name", "customername", "mijoz_nomi", "name", "nomi")
        if not cid:
            continue
        if len(cid) > 64:
            errors.append(GeneralCustomerImportError(row=i, detail="customer_id too long"))
            continue
        cname_val = (cname[:255] if cname else None) or None
        existing = db.query(GeneralCustomerModel).filter(GeneralCustomerModel.customer_id == cid).one_or_none()
        if existing:
            old_data = {"customer_name": existing.customer_name}
            existing.customer_name = cname_val
            log_action(
                db,
                user_id=user.id,
                action=ACTION_UPDATE,
                entity_type="general_customer",
                entity_id=str(existing.id),
                old_data=old_data,
                new_data={"customer_name": cname_val},
                ip_address=get_client_ip(request),
            )
            updated += 1
        else:
            item = GeneralCustomerModel(customer_id=cid, customer_name=cname_val)
            db.add(item)
            db.flush()
            log_action(
                db,
                user_id=user.id,
                action=ACTION_CREATE,
                entity_type="general_customer",
                entity_id=str(item.id),
                new_data={"customer_id": cid},
                ip_address=get_client_ip(request),
            )
            created += 1
    db.commit()
    return GeneralCustomerImportResult(created=created, updated=updated, errors=errors)


@router.post("", response_model=GeneralCustomerOut, status_code=status.HTTP_201_CREATED)
@router.post("/", response_model=GeneralCustomerOut, status_code=status.HTTP_201_CREATED)
async def create_general_customer(
    request: Request,
    payload: GeneralCustomerCreate,
    db: Session = Depends(get_db),
    user=Depends(require_permission("orders:read")),
):
    cid = payload.customer_id.strip()
    existing = db.query(GeneralCustomerModel).filter(GeneralCustomerModel.customer_id == cid).one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="General customer with this customer_id already exists")
    item = GeneralCustomerModel(
        customer_id=cid,
        customer_name=payload.customer_name.strip() if payload.customer_name else None,
    )
    db.add(item)
    log_action(
        db,
        user_id=user.id,
        action=ACTION_CREATE,
        entity_type="general_customer",
        entity_id=str(item.id),
        new_data={"customer_id": item.customer_id},
        ip_address=get_client_ip(request),
    )
    db.commit()
    db.refresh(item)
    return _to_out(item)


@router.put("/{item_id}", response_model=GeneralCustomerOut)
async def update_general_customer(
    request: Request,
    item_id: UUID,
    payload: GeneralCustomerUpdate,
    db: Session = Depends(get_db),
    user=Depends(require_permission("orders:read")),
):
    item = db.query(GeneralCustomerModel).filter(GeneralCustomerModel.id == item_id).one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="General customer not found")
    old_data = {"customer_name": item.customer_name}
    item.customer_name = payload.customer_name.strip() if payload.customer_name else None
    log_action(
        db,
        user_id=user.id,
        action=ACTION_UPDATE,
        entity_type="general_customer",
        entity_id=str(item_id),
        old_data=old_data,
        new_data={"customer_name": item.customer_name},
        ip_address=get_client_ip(request),
    )
    db.commit()
    db.refresh(item)
    return _to_out(item)


@router.delete("/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_general_customer(
    request: Request,
    item_id: UUID,
    db: Session = Depends(get_db),
    user=Depends(require_permission("orders:read")),
):
    item = db.query(GeneralCustomerModel).filter(GeneralCustomerModel.id == item_id).one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="General customer not found")
    log_action(
        db,
        user_id=user.id,
        action=ACTION_DELETE,
        entity_type="general_customer",
        entity_id=str(item_id),
        old_data={"customer_id": item.customer_id},
        ip_address=get_client_ip(request),
    )
    db.delete(item)
    db.commit()
