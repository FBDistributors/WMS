"""VIP customers CRUD: mijoz id, nomi, muddat chegarasi (oy)."""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any, List
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from pydantic import BaseModel, Field
from sqlalchemy import or_
from sqlalchemy.orm import Session, selectinload

from app.auth.deps import require_any_permission, require_permission
from app.db import get_db
from app.models.brand import Brand as BrandModel
from app.models.vip_customer import VipCustomer as VipCustomerModel
from app.models.vip_customer_brand_limit import VipCustomerBrandLimit as VipCustomerBrandLimitModel
from app.services.audit_service import (
    ACTION_CREATE,
    ACTION_DELETE,
    ACTION_UPDATE,
    get_client_ip,
    log_action,
)

router = APIRouter()


class VipBrandLimitOut(BaseModel):
    brand_id: UUID
    min_expiry_months: int


class VipCustomerOut(BaseModel):
    id: UUID
    customer_id: str
    customer_name: str | None
    min_expiry_months: int
    created_at: datetime
    brand_limits: list[VipBrandLimitOut] = Field(default_factory=list)


class VipCustomerCreate(BaseModel):
    customer_id: str = Field(..., min_length=1, max_length=64)
    customer_name: str | None = Field(default=None, max_length=255)
    min_expiry_months: int = Field(..., ge=1, le=60, description="Muddat chegarasi (oy)")


class VipCustomerUpdate(BaseModel):
    customer_name: str | None = Field(default=None, max_length=255)
    min_expiry_months: int | None = Field(default=None, ge=1, le=60)


class VipBrandLimitItemIn(BaseModel):
    brand_id: UUID
    min_expiry_months: int = Field(..., ge=1, le=60)


class VipCustomerImportError(BaseModel):
    row: int
    detail: str


class VipCustomerImportResult(BaseModel):
    created: int
    updated: int
    errors: list[VipCustomerImportError] = Field(default_factory=list)


def _to_out(v: VipCustomerModel) -> VipCustomerOut:
    limits = sorted(
        (
            VipBrandLimitOut(brand_id=b.brand_id, min_expiry_months=b.min_expiry_months)
            for b in (getattr(v, "brand_limits", None) or [])
        ),
        key=lambda x: str(x.brand_id),
    )
    return VipCustomerOut(
        id=v.id,
        customer_id=v.customer_id,
        customer_name=v.customer_name,
        min_expiry_months=v.min_expiry_months,
        created_at=v.created_at,
        brand_limits=limits,
    )


def _norm_header_key(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "_").replace("-", "_")


def _row_dict_from_csv(raw: bytes) -> tuple[list[dict[str, str]], list[VipCustomerImportError]]:
    errors: list[VipCustomerImportError] = []
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        errors.append(VipCustomerImportError(row=0, detail="File must be UTF-8 CSV"))
        return [], errors
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        errors.append(VipCustomerImportError(row=0, detail="CSV has no header row"))
        return [], errors
    rows: list[dict[str, str]] = []
    for i, row in enumerate(reader, start=2):
        clean: dict[str, str] = {}
        for k, v in row.items():
            if k is None:
                continue
            nk = _norm_header_key(str(k))
            if nk:
                clean[nk] = (str(v).strip() if v is not None else "")
        rows.append(clean)
    return rows, errors


def _row_dicts_from_xlsx(raw: bytes) -> tuple[list[dict[str, str]], list[VipCustomerImportError]]:
    errors: list[VipCustomerImportError] = []
    try:
        from openpyxl import load_workbook
    except ImportError:
        errors.append(VipCustomerImportError(row=0, detail="xlsx support not available (openpyxl)"))
        return [], errors
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        errors.append(VipCustomerImportError(row=0, detail=f"Invalid xlsx: {e}"))
        return [], errors
    ws = wb.active
    data_rows: list[tuple[Any, ...]] = []
    for row in ws.iter_rows(values_only=True):
        data_rows.append(row)
    wb.close()
    if not data_rows:
        errors.append(VipCustomerImportError(row=0, detail="Empty sheet"))
        return [], errors
    headers_raw = data_rows[0]
    headers = [_norm_header_key(str(h) if h is not None else "") for h in (headers_raw or ())]
    if not any(h for h in headers):
        errors.append(VipCustomerImportError(row=0, detail="Missing header row"))
        return [], errors
    rows: list[dict[str, str]] = []
    for line_no, vals in enumerate(data_rows[1:], start=2):
        d: dict[str, str] = {}
        for j, h in enumerate(headers):
            if not h:
                continue
            cell = vals[j] if j < len(vals) else None
            d[h] = str(cell).strip() if cell is not None else ""
        rows.append(d)
    return rows, errors


def _get_cell(row: dict[str, str], *keys: str) -> str:
    for k in keys:
        nk = _norm_header_key(k)
        if nk in row:
            return (row.get(nk) or "").strip()
    return ""


def _parse_min_months(raw: str, default: int = 6) -> tuple[int | None, str | None]:
    s = (raw or "").strip()
    if not s:
        return default, None
    try:
        n = int(float(s))
    except ValueError:
        return None, "min_expiry_months must be a number"
    if n < 1 or n > 60:
        return None, "min_expiry_months must be between 1 and 60"
    return n, None


def _vip_with_limits(db: Session, vip_id: UUID) -> VipCustomerModel | None:
    return (
        db.query(VipCustomerModel)
        .options(selectinload(VipCustomerModel.brand_limits))
        .filter(VipCustomerModel.id == vip_id)
        .one_or_none()
    )


@router.get("", response_model=List[VipCustomerOut], summary="List VIP customers")
@router.get("/", response_model=List[VipCustomerOut], summary="List VIP customers")
async def list_vip_customers(
    search: str | None = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    _user=Depends(require_any_permission(["orders:read", "receiving:write", "admin:access"])),
):
    query = db.query(VipCustomerModel)
    if search:
        term = f"%{search.strip()}%"
        query = query.filter(
            or_(
                VipCustomerModel.customer_id.ilike(term),
                VipCustomerModel.customer_name.ilike(term),
            )
        )
    items = (
        query.options(selectinload(VipCustomerModel.brand_limits))
        .order_by(VipCustomerModel.customer_id.asc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return [_to_out(v) for v in items]


_MAX_IMPORT_BYTES = 8 * 1024 * 1024


@router.post(
    "/import",
    response_model=VipCustomerImportResult,
    summary="Import VIP customers from CSV or xlsx (upsert by customer_id)",
)
async def import_vip_customers(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(require_permission("orders:read")),
):
    raw = await file.read()
    if len(raw) > _MAX_IMPORT_BYTES:
        raise HTTPException(status_code=400, detail="File too large (max 8MB)")
    name = (file.filename or "").lower()
    errors: list[VipCustomerImportError] = []
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        rows, parse_err = _row_dicts_from_xlsx(raw)
        errors.extend(parse_err)
    else:
        rows, parse_err = _row_dict_from_csv(raw)
        errors.extend(parse_err)
    if not rows and not errors:
        raise HTTPException(status_code=400, detail="No data rows found")

    created = 0
    updated = 0
    for i, row in enumerate(rows, start=2):
        cid = _get_cell(row, "customer_id", "customerid", "mijoz_id", "id")
        cname = _get_cell(row, "customer_name", "customername", "mijoz_nomi", "name", "nomi")
        months_raw = _get_cell(row, "min_expiry_months", "minexpirymonths", "muddat", "oy")
        if not cid:
            continue
        if len(cid) > 64:
            errors.append(VipCustomerImportError(row=i, detail="customer_id too long"))
            continue
        months, m_err = _parse_min_months(months_raw, default=6)
        if months is None or m_err:
            errors.append(VipCustomerImportError(row=i, detail=m_err or "invalid min_expiry_months"))
            continue
        cname_val = (cname[:255] if cname else None) or None
        existing = db.query(VipCustomerModel).filter(VipCustomerModel.customer_id == cid).one_or_none()
        if existing:
            old_data = {"customer_name": existing.customer_name, "min_expiry_months": existing.min_expiry_months}
            existing.customer_name = cname_val
            existing.min_expiry_months = months
            log_action(
                db,
                user_id=user.id,
                action=ACTION_UPDATE,
                entity_type="vip_customer",
                entity_id=str(existing.id),
                old_data=old_data,
                new_data={"customer_name": cname_val, "min_expiry_months": months},
                ip_address=get_client_ip(request),
            )
            updated += 1
        else:
            vip = VipCustomerModel(
                customer_id=cid,
                customer_name=cname_val,
                min_expiry_months=months,
            )
            db.add(vip)
            db.flush()
            log_action(
                db,
                user_id=user.id,
                action=ACTION_CREATE,
                entity_type="vip_customer",
                entity_id=str(vip.id),
                new_data={"customer_id": cid, "min_expiry_months": months},
                ip_address=get_client_ip(request),
            )
            created += 1
    db.commit()
    return VipCustomerImportResult(created=created, updated=updated, errors=errors)


@router.post("", response_model=VipCustomerOut, status_code=status.HTTP_201_CREATED, summary="Create VIP customer")
@router.post("/", response_model=VipCustomerOut, status_code=status.HTTP_201_CREATED, summary="Create VIP customer")
async def create_vip_customer(
    request: Request,
    payload: VipCustomerCreate,
    db: Session = Depends(get_db),
    user=Depends(require_permission("orders:read")),
):
    existing = db.query(VipCustomerModel).filter(VipCustomerModel.customer_id == payload.customer_id.strip()).one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="VIP customer with this customer_id already exists")
    vip = VipCustomerModel(
        customer_id=payload.customer_id.strip(),
        customer_name=payload.customer_name.strip() if payload.customer_name else None,
        min_expiry_months=payload.min_expiry_months,
    )
    db.add(vip)
    log_action(
        db,
        user_id=user.id,
        action=ACTION_CREATE,
        entity_type="vip_customer",
        entity_id=str(vip.id),
        new_data={"customer_id": vip.customer_id, "min_expiry_months": vip.min_expiry_months},
        ip_address=get_client_ip(request),
    )
    db.commit()
    vip_loaded = _vip_with_limits(db, vip.id)
    return _to_out(vip_loaded) if vip_loaded else _to_out(vip)


@router.put("/{vip_id}", response_model=VipCustomerOut, summary="Update VIP customer")
async def update_vip_customer(
    request: Request,
    vip_id: UUID,
    payload: VipCustomerUpdate,
    db: Session = Depends(get_db),
    user=Depends(require_permission("orders:read")),
):
    vip = db.query(VipCustomerModel).filter(VipCustomerModel.id == vip_id).one_or_none()
    if not vip:
        raise HTTPException(status_code=404, detail="VIP customer not found")
    old_data = {"customer_name": vip.customer_name, "min_expiry_months": vip.min_expiry_months}
    if payload.customer_name is not None:
        vip.customer_name = payload.customer_name.strip() or None
    if payload.min_expiry_months is not None:
        vip.min_expiry_months = payload.min_expiry_months
    log_action(
        db,
        user_id=user.id,
        action=ACTION_UPDATE,
        entity_type="vip_customer",
        entity_id=str(vip_id),
        old_data=old_data,
        new_data={"customer_name": vip.customer_name, "min_expiry_months": vip.min_expiry_months},
        ip_address=get_client_ip(request),
    )
    db.commit()
    vip_loaded = _vip_with_limits(db, vip_id)
    return _to_out(vip_loaded) if vip_loaded else _to_out(vip)


@router.put(
    "/{vip_id}/brand-limits",
    response_model=VipCustomerOut,
    summary="Replace VIP per-brand expiry limits",
)
async def replace_vip_brand_limits(
    request: Request,
    vip_id: UUID,
    payload: List[VipBrandLimitItemIn],
    db: Session = Depends(get_db),
    user=Depends(require_permission("orders:read")),
):
    vip = db.query(VipCustomerModel).filter(VipCustomerModel.id == vip_id).one_or_none()
    if not vip:
        raise HTTPException(status_code=404, detail="VIP customer not found")
    seen_brands: set[UUID] = set()
    for item in payload:
        if item.brand_id in seen_brands:
            raise HTTPException(status_code=400, detail="Duplicate brand_id in payload")
        seen_brands.add(item.brand_id)
        brand = db.query(BrandModel).filter(BrandModel.id == item.brand_id).one_or_none()
        if not brand:
            raise HTTPException(status_code=400, detail=f"Brand not found: {item.brand_id}")

    old_limits = [
        {"brand_id": str(b.brand_id), "min_expiry_months": b.min_expiry_months}
        for b in db.query(VipCustomerBrandLimitModel).filter(VipCustomerBrandLimitModel.vip_customer_id == vip_id).all()
    ]
    db.query(VipCustomerBrandLimitModel).filter(VipCustomerBrandLimitModel.vip_customer_id == vip_id).delete(
        synchronize_session=False
    )
    for item in payload:
        db.add(
            VipCustomerBrandLimitModel(
                vip_customer_id=vip_id,
                brand_id=item.brand_id,
                min_expiry_months=item.min_expiry_months,
            )
        )
    new_limits = [{"brand_id": str(i.brand_id), "min_expiry_months": i.min_expiry_months} for i in payload]
    log_action(
        db,
        user_id=user.id,
        action=ACTION_UPDATE,
        entity_type="vip_customer",
        entity_id=str(vip_id),
        old_data={"brand_limits": old_limits},
        new_data={"brand_limits": new_limits},
        ip_address=get_client_ip(request),
    )
    db.commit()
    vip_loaded = _vip_with_limits(db, vip_id)
    if not vip_loaded:
        raise HTTPException(status_code=404, detail="VIP customer not found")
    return _to_out(vip_loaded)


@router.delete("/{vip_id}", status_code=status.HTTP_204_NO_CONTENT, summary="Delete VIP customer")
async def delete_vip_customer(
    request: Request,
    vip_id: UUID,
    db: Session = Depends(get_db),
    user=Depends(require_permission("orders:read")),
):
    vip = db.query(VipCustomerModel).filter(VipCustomerModel.id == vip_id).one_or_none()
    if not vip:
        raise HTTPException(status_code=404, detail="VIP customer not found")
    log_action(
        db,
        user_id=user.id,
        action=ACTION_DELETE,
        entity_type="vip_customer",
        entity_id=str(vip_id),
        old_data={"customer_id": vip.customer_id},
        ip_address=get_client_ip(request),
    )
    db.delete(vip)
    db.commit()
